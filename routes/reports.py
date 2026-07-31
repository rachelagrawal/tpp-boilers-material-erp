from flask import (
    Blueprint,
    render_template,
    send_file
)

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)

from io import BytesIO
from datetime import datetime

from models import (
    Inventory,
    Item
)

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.units import inch

from datetime import datetime

from models import (
    Inventory,
    Item,
    PurchaseOrder,
    Supplier,
    GRN,
    StockIssue
)


report_bp = Blueprint("reports", __name__)

@report_bp.route("/reports/inventory")
def inventory_report():

    inventory = Inventory.query.all()

    return render_template(
        "inventory_report.html",
        inventory=inventory
    )

@report_bp.route("/reports/inventory/excel")
def inventory_report_excel():

    inventory = Inventory.query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory Report"

    # ==========================================
    # Title
    # ==========================================

    sheet["A1"] = "TPP Boilers Inventory Report"

    sheet["A1"].font = Font(
        bold=True,
        size=18
    )

    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    sheet.merge_cells("A1:D1")

    # ==========================================
    # Date Generated
    # ==========================================

    sheet["A2"] = (
        "Generated on: "
        + datetime.now().strftime("%d-%b-%Y %I:%M %p")
    )

    sheet["A2"].font = Font(
        italic=True,
        size=11
    )

    # ==========================================
    # Headers
    # ==========================================

    headers = [
        "Item Code",
        "Material",
        "Current Stock",
        "Minimum Stock"
    ]

    sheet.append([])
    sheet.append(headers)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for cell in sheet[4]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    # ==========================================
    # Inventory Data
    # ==========================================

    low_stock_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    healthy_fill = PatternFill(
        start_color="E2F0D9",
        end_color="E2F0D9",
        fill_type="solid"
    )

    for stock in inventory:

        sheet.append([
            stock.item.item_code,
            stock.item.material_name,
            stock.current_stock,
            stock.item.min_stock
        ])

        row = sheet.max_row

        if (
            stock.item.min_stock
            and stock.current_stock < stock.item.min_stock
        ):

            fill = low_stock_fill

        else:

            fill = healthy_fill

        for cell in sheet[row]:

            cell.fill = fill

        sheet[f"C{row}"].alignment = Alignment(
            horizontal="center"
        )

        sheet[f"D{row}"].alignment = Alignment(
            horizontal="center"
        )

    # ==========================================
    # Summary
    # ==========================================

    sheet.append([])

    sheet.append([
        "",
        "",
        "Total Items",
        len(inventory)
    ])

    last_row = sheet.max_row

    sheet[f"C{last_row}"].font = Font(
        bold=True
    )

    sheet[f"D{last_row}"].font = Font(
        bold=True
    )

    # ==========================================
    # Column Widths
    # ==========================================

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 18

    # ==========================================
    # Freeze Header
    # ==========================================

    sheet.freeze_panes = "A5"

    # ==========================================
    # Export
    # ==========================================

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        download_name="inventory_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@report_bp.route("/reports")
def reports():

    return render_template("reports.html")


@report_bp.route("/reports/inventory/pdf")
def inventory_report_pdf():

    inventory = Inventory.query.all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elements = []

    # ==========================================
    # Company Name
    # ==========================================

    company = Paragraph(
        "<font size=22><b>TPP BOILERS</b></font>",
        styles["Title"]
    )

    elements.append(company)

    # ==========================================

    title = Paragraph(
        "<font size=16><b>Inventory Report</b></font>",
        styles["Heading2"]
    )

    elements.append(title)

    elements.append(
        Paragraph(
            f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1, 0.3 * inch))

    # ==========================================
    # Table Data
    # ==========================================

    data = [[
        "Item Code",
        "Material",
        "Current",
        "Minimum",
        "Status"
    ]]

    low_stock_count = 0

    for stock in inventory:

        if (
            stock.item.min_stock
            and stock.current_stock < stock.item.min_stock
        ):

            status = "LOW"

            low_stock_count += 1

        else:

            status = "OK"

        data.append([

            stock.item.item_code,

            stock.item.material_name,

            str(stock.current_stock),

            str(stock.item.min_stock),

            status

        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")),

        ("TEXTCOLOR", (0,0), (-1,0), colors.white),

        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

        ("FONTSIZE", (0,0), (-1,0), 11),

        ("BOTTOMPADDING", (0,0), (-1,0), 10),

        ("GRID", (0,0), (-1,-1), 1, colors.grey),

        ("ALIGN", (2,1), (4,-1), "CENTER"),

        ("BACKGROUND", (0,1), (-1,-1), colors.beige)

    ]))

    # ==========================================
    # Highlight Low Stock Rows
    # ==========================================

    for row in range(1, len(data)):

        if data[row][4] == "LOW":

            table.setStyle(TableStyle([

                ("BACKGROUND", (0,row), (-1,row),
                 colors.lightcoral)

            ]))

    elements.append(table)

    elements.append(Spacer(1, 0.3 * inch))

    # ==========================================
    # Summary
    # ==========================================

    summary = Paragraph(

        f"""
        <b>Total Inventory Items:</b> {len(inventory)}
        <br/>
        <b>Low Stock Items:</b> {low_stock_count}
        """,

        styles["Heading3"]

    )

    elements.append(summary)

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        download_name="inventory_report.pdf",
        as_attachment=True,
        mimetype="application/pdf"
    )


@report_bp.route("/reports/purchase-orders")
def purchase_order_report():

    purchase_orders = PurchaseOrder.query.all()

    return render_template(
        "purchase_order_report.html",
        purchase_orders=purchase_orders
    )

@report_bp.route("/reports/suppliers")
def supplier_report():

    suppliers = Supplier.query.all()

    return render_template(
        "supplier_report.html",
        suppliers=suppliers
    )

@report_bp.route("/reports/grns")
def grn_report():

    grns = GRN.query.all()

    return render_template(
        "grn_report.html",
        grns=grns
    )

@report_bp.route("/reports/stock-issues")
def stock_issue_report():

    stock_issues = StockIssue.query.all()

    return render_template(
        "stock_issue_report.html",
        stock_issues=stock_issues
    )

@report_bp.route("/reports/purchase-orders/excel")
def purchase_order_report_excel():

    purchase_orders = PurchaseOrder.query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Purchase Orders"

    sheet["A1"] = "TPP Boilers Purchase Order Report"

    sheet["A1"].font = Font(
        bold=True,
        size=18
    )

    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A1:D1")

    sheet["A2"] = (
        "Generated on: "
        + datetime.now().strftime("%d-%b-%Y %I:%M %p")
    )

    sheet.append([])

    headers = [
        "PO Number",
        "Supplier",
        "Order Date",
        "Status"
    ]

    sheet.append(headers)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for cell in sheet[4]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(horizontal="center")

    for po in purchase_orders:

        sheet.append([
            po.po_number,
            po.supplier.supplier_name,
            str(po.order_date),
            po.status
        ])

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 18

    sheet.freeze_panes = "A5"

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        download_name="purchase_order_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@report_bp.route("/reports/purchase-orders/pdf")
def purchase_order_report_pdf():

    purchase_orders = PurchaseOrder.query.all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<font size=22><b>TPP BOILERS</b></font>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "<font size=16><b>Purchase Order Report</b></font>",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1,0.3*inch))

    data = [[
        "PO No",
        "Supplier",
        "Date",
        "Status"
    ]]

    for po in purchase_orders:

        data.append([
            po.po_number,
            po.supplier.supplier_name,
            str(po.order_date),
            po.status
        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("GRID",(0,0),(-1,-1),1,colors.grey),
        ("BACKGROUND",(0,1),(-1,-1),colors.beige),
        ("ALIGN",(0,0),(-1,-1),"CENTER")

    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        download_name="purchase_order_report.pdf",
        as_attachment=True,
        mimetype="application/pdf"
    )

@report_bp.route("/reports/suppliers/excel")
def supplier_report_excel():

    suppliers = Supplier.query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Suppliers"

    sheet["A1"] = "TPP Boilers Supplier Report"

    sheet["A1"].font = Font(
        bold=True,
        size=18
    )

    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A1:E1")

    sheet["A2"] = (
        "Generated on: "
        + datetime.now().strftime("%d-%b-%Y %I:%M %p")
    )

    sheet.append([])

    headers = [
        "Supplier Code",
        "Supplier Name",
        "Contact Person",
        "Phone",
        "Email"
    ]

    sheet.append(headers)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for cell in sheet[4]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(horizontal="center")

    for supplier in suppliers:

        sheet.append([
            supplier.supplier_code,
            supplier.supplier_name,
            supplier.contact_person,
            supplier.phone,
            supplier.email
        ])

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 35

    sheet.freeze_panes = "A5"

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        download_name="supplier_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@report_bp.route("/reports/suppliers/pdf")
def supplier_report_pdf():

    suppliers = Supplier.query.all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<font size=22><b>TPP BOILERS</b></font>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "<font size=16><b>Supplier Report</b></font>",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1,0.3*inch))

    data = [[
        "Code",
        "Supplier",
        "Contact",
        "Phone",
        "Email"
    ]]

    for supplier in suppliers:

        data.append([
            supplier.supplier_code,
            supplier.supplier_name,
            supplier.contact_person,
            supplier.phone,
            supplier.email
        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("GRID",(0,0),(-1,-1),1,colors.grey),
        ("BACKGROUND",(0,1),(-1,-1),colors.beige),
        ("ALIGN",(0,0),(-1,-1),"CENTER")

    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        download_name="supplier_report.pdf",
        as_attachment=True,
        mimetype="application/pdf"
    )

@report_bp.route("/reports/grns/excel")
def grn_report_excel():

    grns = GRN.query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GRN Report"

    sheet["A1"] = "TPP Boilers GRN Report"

    sheet["A1"].font = Font(
        bold=True,
        size=18
    )

    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A1:D1")

    sheet["A2"] = (
        "Generated on: "
        + datetime.now().strftime("%d-%b-%Y %I:%M %p")
    )

    sheet.append([])

    headers = [
        "GRN Number",
        "PO Number",
        "Supplier",
        "Received Date"
    ]

    sheet.append(headers)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for cell in sheet[4]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(horizontal="center")

    for grn in grns:

        sheet.append([
            grn.grn_number,
            grn.purchase_order.po_number,
            grn.purchase_order.supplier.supplier_name,
            str(grn.received_date)
        ])

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 35
    sheet.column_dimensions["D"].width = 20

    sheet.freeze_panes = "A5"

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        download_name="grn_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@report_bp.route("/reports/grns/pdf")
def grn_report_pdf():

    grns = GRN.query.all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<font size=22><b>TPP BOILERS</b></font>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "<font size=16><b>GRN Report</b></font>",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1, 0.3 * inch))

    data = [[
        "GRN No",
        "PO No",
        "Supplier",
        "Received Date"
    ]]

    for grn in grns:

        data.append([
            grn.grn_number,
            grn.purchase_order.po_number,
            grn.purchase_order.supplier.supplier_name,
            str(grn.received_date)
        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 1, colors.grey),
        ("BACKGROUND", (0,1), (-1,-1), colors.beige),
        ("ALIGN", (0,0), (-1,-1), "CENTER")

    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        download_name="grn_report.pdf",
        as_attachment=True,
        mimetype="application/pdf"
    )

@report_bp.route("/reports/stock-issues/excel")
def stock_issue_report_excel():

    stock_issues = StockIssue.query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Stock Issues"

    sheet["A1"] = "TPP Boilers Stock Issue Report"

    sheet["A1"].font = Font(
        bold=True,
        size=18
    )

    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A1:D1")

    sheet["A2"] = (
        "Generated on: "
        + datetime.now().strftime("%d-%b-%Y %I:%M %p")
    )

    sheet.append([])

    headers = [
        "Item",
        "Quantity",
        "Issue Date",
        "Department"
    ]

    sheet.append(headers)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for cell in sheet[4]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(horizontal="center")

    for issue in stock_issues:

        sheet.append([
            issue.item.material_name,
            issue.quantity,
            str(issue.issue_date),
            issue.department
        ])

    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 25

    sheet.freeze_panes = "A5"

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        download_name="stock_issue_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@report_bp.route("/reports/stock-issues/pdf")
def stock_issue_report_pdf():

    stock_issues = StockIssue.query.all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<font size=22><b>TPP BOILERS</b></font>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "<font size=16><b>Stock Issue Report</b></font>",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1, 0.3 * inch))

    data = [[
        "Item",
        "Quantity",
        "Issue Date",
        "Department"
    ]]

    for issue in stock_issues:

        data.append([
            issue.item.material_name,
            str(issue.quantity),
            str(issue.issue_date),
            issue.department
        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 1, colors.grey),
        ("BACKGROUND", (0,1), (-1,-1), colors.beige),
        ("ALIGN", (0,0), (-1,-1), "CENTER")

    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        download_name="stock_issue_report.pdf",
        as_attachment=True,
        mimetype="application/pdf"
    )